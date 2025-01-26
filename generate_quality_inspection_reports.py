import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import os

def generate_reports(source_file='examples/example_source_data.xls', 
                    template_file='examples/example_template.xlsx',
                    output_file='quality_inspection_report.xlsx'):
    """
    生成质量检验报告。

    参数:
        source_file (str): 源数据文件路径
        template_file (str): 模板文件路径
        output_file (str): 输出文件路径
    """
    # 检查文件是否存在
    if not os.path.exists(source_file):
        raise FileNotFoundError(f"源数据文件不存在: {source_file}")
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"模板文件不存在: {template_file}")

    # 读取第一个文件
    header_df = pd.read_excel(source_file, sheet_name='HEADER')

    # 读取Dimension表，跳过前12行，然后使用第13行作为列名
    dimension_df = pd.read_excel(source_file, sheet_name='Dimension', skiprows=12)
    # 使用第一行作为列名
    dimension_df.columns = dimension_df.iloc[0]
    # 删除第一行（现在已经作为列名）并重置索引
    dimension_df = dimension_df.iloc[1:].reset_index(drop=True)

    # 读取Sand表的数据
    sand_df = pd.read_excel(source_file, sheet_name='Sand', header=None)

    # 读取第二个文件
    wb = load_workbook(template_file)
    wacker_sheet = wb['WACKER']

    # 获取Sales Order Quantity和Quality Assured By
    sales_order_quantity = header_df.iloc[5, 2]  # Sales Order Quantity位置
    quality_assured_by = header_df.iloc[3, 7]    # Quality Assured By在第4行最后一列

    # 定义元素和行号的对应关系
    element_row_mapping = {
        'Al': 9,   # 硅柏_石英坩埚_QC530HS_201410_V3B-CN_Al
        'Ca': 10,  # 硅柏_石英坩埚_QC530HS_201410_V3B-CN_Ca
        'Cu': 11,  # 硅柏_石英坩埚_QC530HS_201410_V3B-CN_Cu
        'Fe': 12,  # 硅柏_石英坩埚_QC530HS_201410_V3B-CN_Fe
        'K': 13,   # 硅柏_石英坩埚_QC530HS_201410_V3B-CN_K
        'Li': 14,  # 硅柏_石英坩埚_QC530HS_201410_V3B-CN_Li
        'Mg': 15,  # 硅柏_石英坩埚_QC530HS_201410_V3B-CN_Mg
        'Mn': 16,  # 硅柏_石英坩埚_QC530HS_201410_V3B-CN_Mn
        'Na': 17,  # 硅柏_石英坩埚_QC530HS_201410_V3B-CN_Na
        'Ti': 18,  # 硅柏_石英坩埚_QC530HS_201410_V3B-CN_Ti
        'Zr': 19   # 硅柏_石英坩埚_QC530HS_201410_V3B-CN_Zr
    }

    # 定义元素在Sand表中的列索引
    element_col_mapping = {
        'Al': 4,   # 第5列
        'Ca': 5,   # 第6列
        'Cu': 6,   # 第7列
        'Fe': 7,   # 第8列
        'K': 8,    # 第9列
        'Li': 9,   # 第10列
        'Mg': 10,  # 第11列
        'Mn': 11,  # 第12列
        'Na': 12,  # 第13列
        'Ti': 13,  # 第14列
        'Zr': 14   # 第15列
    }

    # 遍历Dimension表格中的每个Customer ID
    for index, row in dimension_df.iterrows():
        customer_id = row['Customer ID']  # 现在这个列名应该是正确的了
        inspection_date = pd.to_datetime(row['Inspection Date']).strftime('%Y-%m-%d')  # 格式化日期
        
        # 创建新的工作表
        new_sheet = wb.create_sheet(title=str(customer_id))
        
        # 复制WACKER表格的内容到新工作表（这样会保持原有的客户名称）
        for row_wacker in wacker_sheet.iter_rows(values_only=True):
            new_sheet.append(row_wacker)
        
        # 填充数据（不再覆盖客户名称）
        new_sheet['B3'] = str(sales_order_quantity) + ' PCS'  # Number+Unit/数量+单位
        new_sheet['B4'] = customer_id  # Batch reference/批号
        new_sheet['D4'] = inspection_date  # Date of issue/报告日期
        new_sheet['B5'] = inspection_date  # Production date/生产日期
        new_sheet['D5'] = (datetime.strptime(inspection_date, '%Y-%m-%d') + timedelta(days=730)).strftime('%Y-%m-%d')  # Expiring date/失效日期
        
        # 从sand表中获取当前customer_id的数据
        sand_rows = sand_df[sand_df[2] == customer_id]  # 使用第3列（索引2）作为Crucible ID
        if not sand_rows.empty:
            sand_row = sand_rows.iloc[0]
            
            # 填充元素数据
            for element, target_row in element_row_mapping.items():
                source_col = element_col_mapping[element]
                new_sheet[f'D{target_row}'] = sand_row[source_col]
        
        # 填充Analysis result/分析结果
        # 保持原有的测试项目名称，只更新分析结果列
        for i in range(20, 29):
            if i == 20:
                new_sheet[f'D{i}'] = row['OD1']  # 外径1
            elif i == 21:
                new_sheet[f'D{i}'] = row['OD2']  # 外径2
            elif i == 22:
                new_sheet[f'D{i}'] = row['OD3']  # 外径3
            elif i == 23:
                new_sheet[f'D{i}'] = row['Height']  # 高度
            elif i == 24:
                new_sheet[f'D{i}'] = row['Wall11']  # 壁厚11
            elif i == 25:
                new_sheet[f'D{i}'] = row['Wall12']  # 壁厚12
            elif i == 26:
                new_sheet[f'D{i}'] = row['Wall13']  # 壁厚13
            elif i == 27:
                new_sheet[f'D{i}'] = row['Wall2']  # 壁厚2
            elif i == 28:
                new_sheet[f'D{i}'] = row['Wall3']  # 壁厚3
        
        # 保持"批准人："文本，并在其后添加名字
        new_sheet['D29'] = f"批准人：{quality_assured_by}"

    # 保存修改后的文件
    wb.save(output_file)
    print(f"报告已生成: {output_file}")

if __name__ == "__main__":
    generate_reports()